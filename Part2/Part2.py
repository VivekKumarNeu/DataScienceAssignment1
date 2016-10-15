import requests
import bs4
import PyPDF2
import re
import os
import csv
import openpyxl
import shutil

# function to find the quarter of a given month
def findQuarter(month):
    lowerMonth=month.lower()
    if(lowerMonth=='mar'):
        return 'Quarter1'
    if (lowerMonth == 'jun'):
        return 'Quarter2'
    if (lowerMonth == 'sep'):
        return 'Quarter3'
    if (lowerMonth == 'dec'):
        return 'Quarter4'

# function to add a sheet to the workbook.
# Cleans up data using regular expressions
def addSheet(pageNum,dirName,fileName,wb):
    sheet = wb.create_sheet()
    sheet.title = 'Page_'+str(pageNum)
    exampleFile = open(os.path.join('.',str(dirName), str(fileName)))
    exampleReader = csv.reader(exampleFile)
    workRow = 1
    mainRow=1
    for row in exampleReader:
        workCol = 1
        for column in range(len(row)):
            cell = row[column]
            if column % 5 == 0 and column!=0:
                index = re.search('[A-Za-z]+', cell)
                if (index):
                    pref=cell[:index.start()]
                    suf=cell[index.start():]
                    if(pref!=''):
                        sheet.cell(row=workRow, column=workCol).value = cell[:index.start()]
                        workRow = workRow + 1
                        workCol = 1
                        sheet.cell(row=workRow, column=workCol).value = cell[index.start():]
                    else:
                        if(mainRow==1):
                            workRow = workRow + 1
                            workCol = 1
                            sheet.cell(row=workRow, column=workCol).value = cell
                        else:
                            sheet.cell(row=workRow, column=workCol).value = cell

                else:
                    sheet.cell(row=workRow, column=workCol).value = cell
            else:
                if mainRow not in [1,2] and column==0 and row[column]=='':
                    workCol=workCol-1
                else:
                    sheet.cell(row=workRow, column=workCol).value = cell
            workCol = workCol + 1
        workRow = workRow + 1
        mainRow=mainRow+1

# Initial Set up
res=requests.get('https://www.ffiec.gov/nicpubweb/content/BHCPRRPT/BHCPR_Peer.htm')
url='https://www.ffiec.gov/nicpubweb/content/BHCPRRPT/'
res.raise_for_status()
noScratchsoup=bs4.BeautifulSoup(res.text)
div=noScratchsoup.select('div.contentfull table a')

# Get the Pdf files, by looping through each anchor tag
for link in div:
    txt=link.get('href')
    if(txt.find('PeerGroup_1')!=-1):
        r=requests.get(url+txt)
        fileName=txt[txt.find('PeerGroup'):len(txt)]
        with open(fileName,"wb") as code:
            code.write(r.content)

# Store the names of all the pdf's that got downloaded.
pdfFiles = []
for filename in os.listdir('.'):
    if filename.endswith('.pdf'):
        pdfFiles.append(filename)

# loop through each pdf and convert it into CSV. CSV are created per page and placed into directory
for filename in pdfFiles:
    quarter = findQuarter(filename[12:15])
    temp = filename[-8:]
    year = temp[0:4]
    directoryName='Peer1_'+year+'_'+quarter
    os.makedirs(directoryName, exist_ok=True)
    pdfFileObj = open(filename, 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    pageText = ''
    for i in range(0, pdfReader.getNumPages()):
        pageObj = pdfReader.getPage(i)
        pageText = pageObj.extractText()
        with open(os.path.join(directoryName,"Peer1_"+year+"_"+quarter+"_Page_"+str(i)+".csv"), "w") as code:
            code.write(re.sub('[^\S\r\n]{3,}', ',', pageText))

# Convert the CSV in individual directory, add each CSV as a sheet into the Workbook.
for dirname in os.listdir('.'):
    wb = openpyxl.Workbook()
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    if dirname.startswith('Peer1'):
        lst=os.listdir(os.path.join('.',dirname))
        ordered_files = sorted(lst, key=lambda x: (int(re.sub('\D', '', x)), x)) # Solution found in stackover flow, to get ordered folders
        for filename in ordered_files:
            i=0
            if filename.endswith('.csv'):
                addSheet(i,dirname,filename,wb)
                i=i+1
        wb.save(str(dirname)+'.xlsx')
        shutil.rmtree(os.path.join('.',dirname))